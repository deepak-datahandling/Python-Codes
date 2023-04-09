# Databricks notebook source
dbutils.widgets.text("Host","","")
dbutils.widgets.text("User","","")
dbutils.widgets.text("Password","","")
dbutils.widgets.text("DatabaseName","","")
dbutils.widgets.text("TableName","","")
dbutils.widgets.text("ADLS_Path","","")
dbutils.widgets.text("TargetTable","","")


Host = dbutils.widgets.get("Host")
User = dbutils.widgets.get("User")
Password = dbutils.widgets.get("Password")
DatabaseName = dbutils.widgets.get("DatabaseName")
TableName = dbutils.widgets.get("TableName")
ADLS_Path = dbutils.widgets.get("ADLS_Path")
Target_Table=dbutils.widgets.get("TargetTable")

tble_name=DatabaseName + '.' + TableName

print("Host: "+Host)
print("User: "+User)
print("DatabaseName: "+DatabaseName)
print("TableName: "+TableName)
print("ADLS_Path: "+ADLS_Path)
print("TargetTable: "+Target_Table)

#-------------------------------------------------------------------IMPORTING REQUIRED LIBRARIES---------------------------------------------------------

#teradata
import teradatasql
import pandas as pd
import pyspark as ps
import re

#parquet
from pyspark.sql.functions import *

#directory 
import os

#excel file handling
from openpyxl import *
from shutil import copyfile
from openpyxl.styles import Font,Fill,Alignment,Border,Side,PatterFill
from openpyxl.worksheet.dimensions import ColumnDimension
import openpyxl as op

#-------------------------------------------------------------------TERADATA DATABASE HANDLING-------------------------------------------------------------
#Teradata Credentials
def host_check(host):
    if(host != '<host name>'):
        raise Exception("Invalid Host Name")
try:
    h=Host
    u=User
    ps=Password
    host_check(h)
    con=teradatasql.connect(host=h,user=u,password=ps)
#Connecting with Teradata
    cursor=con.cursor()
except Exception as h:
    raise ConnectionError(h)

 #Getting Column Names of a table
col_details=f'SELECT ColumnName FROM dbc.columns WHERE DatabaseName =\'{DatabaseName}\' and TableName = \'{TableName}\' ORDER BY Columnid'
cursor.execute(col_details)
col_res=cursor.fetchall()

def col_list():
    lk=[]
    for i in range (0,len(col_res)):
        col=re.sub('[^\W]','',str(col_res[i]))
        lk.append(col)
    return lk

columns=col_list()

#Declaring Stats Name into dictionary for final output
dict_list=['count','min','max','Null_Count']
dict1={'count':[],'min':[],'max':[],'Null_Count':[]}

#Primary Key Field
pk_query=f'''SEL ColumnName FROM dbc.indicesV where TableName like '%{TableName[1:]}' order by ColumnPosition'''

cursor.execute(pk_query)
pk_res=cursor.fetchall()

#Order By Fields
def orderby():
	pk_res1=[]
	filter_fields=''
	for j in pk_res:
		for k in j:
			if k not in pk_res1:
				pk_res1.append(k)
	for l in pk_res1:
		if (pk_res1[-1]==l):
			filter_fields+=l
		else:
			filter_fields+=l
			filter_fields=','
	return filter_fields

order_by_fields=orderby()

#Sample Records

sample_query=f'SEL TOP 5* FROM {DatabaseName}.{TableName} ORDER BY {order_by_fields}'

cursor.execute(sample_query)
sample_res=cursor.fetchall()

sample_tera_df=pd.Dataframe(sample_res,columns=columns)

#Collecting Stats

for j in range (0,len(columns)):
	count_query=f'SEL COUNT({columns[j]}) FROM {DatabaseName}.{TableName}'
	max_query=f'SEL MAX({columns[j]}) FROM {DatabaseName}.{TableName}'
	min_query=f'SEL MIN({columns[j]}) FROM {DatabaseName}.{TableName}'
	null_count_query=f'SEL COUNT(*)-COUNT({columns[j]}) as Null_Count FROM {DatabaseName}.{TableName}'

	cursor.execute(count_query)
	count_res=cursor.fetchall()[0][0]

	cursor.execute(max_query)
	max_res=cursor.fetchall()[0][0]

	cursor.execute(min_query)
	min_res=cursor.fetchall()[0][0]

	cursor.execute(null_count_query)
	null_res=cursor.fetchall()[0][0]

	t=[count_res,min_res,max_res,null_res]
	counter=0
	for k in t:
		f_data=str(k).rstrip()
		dict1[dict_list[counter]].append(f_data)
		counter=counter+1

con.close()

tera_out=pd.Dataframe(dict1)

#index Renaming
for i in range(0,len(column)):
    tera_out=tera_out.rename(index={i:column[i]})

#Converting Stats into String for comparison
tera_out=tera_out.astype(str).convert_dtypes(infer_objects=True)

print(tera_out)


#------------------------------------------------------------------PARQUET FILE HANDLING--------------------------------------------------------

#Reading a Parquet File
full_df1=spark.read.format("Parquet")\
			.option("recursiveFileLookup","true")\
			.load(ADLS_Path)

#Convert Default Databricks CST Timestamp to UTC Timestamp Format
full_df2=full_df1.select([
	from_utc_timestamp(col(l), 'UTC+5').alias(l) if (k == 'timestamp' or k== 'date') else l for l,k in full_df1.dtypes
	])

#Convert Timestamp fields into String for comparison
full_df3=full_df2.select([
	col(l).cast('string') if (k == 'timestamp' or k== 'date') else l for l,k in full_df1.dtypes 
	])

#COMMAND-------------

#Getting Summary Stats of Parquet File

parq_desc=full_df3.describe()

display(parq_desc)

#COMMAND-------------

#Removing Unwanted White Spaces
for c in parq_desc.columns:
	parq_desc=parq_desc.withColumn(c,rtrim(parq_desc[c]))

#Calculating Null Count for Parquet File Fields
null_c=full_df3.select([(count(when((upper(col(c))=='NULL') | \
								isnan(c) | \
								(upper(col(c))=='NONE') | \
								(col(c)==''),c))
								).alias(c) for c in full_df3.columns])

#COMMAND-------------

#converting Null DF into Pandas

null_c=null_c.toPandas()

#COMMAND-------------

parq_pd=parq_desc.toPandas()

parq_pd=parq_pd.set_index("summary").T
null_c=null_c.T
null_count=null_c.rename(columns={0:'Null_Count'})

req_stats=['count','min','max']

temp=[]
for j in req_stats:
	temp.append(parq_pd[j])

parq_temp=pd.Dataframe(temp)
parq_temp=parq_temp.T

parq_out=parq_temp.join(null_count)

#droping the unwanted stats for proper comparison
parq_out=parq_out.drop('AZURE_LOAD_DTTM')

#Converting Stats into String for comparison
parq_out=parq_out.astype(str).convert_dtypes(infer_objects=True)

#Sample Records

full_df3.createOrReplaceTempView(Target_Table)

query=f'SELECT * FROM {Target_Table} ORDER BY {order_by_fields} LIMIT 5'

temp_out=spark.sql(query)

sample_parq_df=temp_out.toPandas()

#-------------------------------------------------------------STATS COMPARSION-----------------------------------------------------------------------------

comp_out=(tera_out == parq_out)

print(comp_out)


#-----------------------------------------------------------STORING ALL THE RESULT IN EXCEL FILE--------------------------------------------------------

#Creating Temp Directory and make it as Working directory to store excel report and mail body contents

#for mail body contents
os.chdir('dbfs/teradata')
o_dir=os.getcwd()
print('Present Working Directory: ',o_dir)

#Creating empty excel for report making
# wb=op.Workbook()
# wb.save('/tmp/data_validation.xlsx')

#Inserting Validation Results into Excel File
with pd.ExcelWriter('/tmp/data_validation.xlsx',engine='openpyxl',mode='a',if_sheet_exists='overlay') as write:
	tera_out.to_excel(write,sheet_name=Target_Table,startrow=2)
	parq_out.to_excel(write,sheet_name=Target_Table,startrow=len(columns)+6)
	comp_out.to_excel(write,sheet_name=Target_Table,startrow=(2*len(columns)+10))
	sample_tera_df.to_excel(write,sheet_name=Target_Table,startcol=8,startrow=4)
	sample_parq_df.to_excel(write,sheet_name=Target_Table,startcol=8,startrow=len(columns)+8)


excel_handling=load_workbook('/tmp/data_validation.xlsx')

#Removing Empty Sheet
if('Sheet' in excel_handling.sheetnames):
	excel_handling.remove(excel_handling['Sheet'])

#Inserting the leftout details like headings,highlights etc.,
if(Target_Table in excel_handling.sheetnames):
	act=excel_handling.active=excel_handling[Target_Table]
#Cell Handling
	act['G1']='Teradata Table'
	act['G1'].font=Font(size=14,bold=True)
	act.merge_cells(start_row=3, start_column=10, end_row=3, end_column=100)
	act['A2']='Summary Stats'
	act['A2'].font=Font(size=13,bold=True)
	act['J2']='Query:'
	act['J2'].font=Font(size=13,bold=True)
	act['J3']=sample_query
	act['G'+str(len(columns)+5)]='Parquet File'
	act['G'+str(len(columns)+5)].font=Font(size=14,bold=True)
	act['A'+str(len(columns)+6)]='Summary Stats'
	act['A'+str(len(columns)+6)].font=Font(size=13,bold=True)
	act['J'+str(len(columns)+6)]='Query:'
	act['J'+str(len(columns)+6)].font=Font(size=13,bold=True)
	act.merge_cells(start_row=len(columns)+7, start_column=10, end_row=len(columns)+7, end_column=100)
	act['J'+str(len(columns)+7)]=query
	act['G'+str(2*len(columns)+9)]='Comparison'
	act['G'+str(2*len(columns)+9)].font=Font(size=14,bold=True)

#Cell Styles
	thin_border = Border(lef=Side(style='thin'),
						right=Side(style='thin'),
						top=Side(style='thin'),
						bottom=Side(style='thin'))

	range_tera_stats=act['A3':'E'+str(len(columns)+3)]
	for cell in range_tera_stats:
		for x in cell:
			x.border=thin_border

	range_parq_stats=act['A'+str(len(columns)+7) : 'E'+str(2*len(columns)+7)]
	for cell in range_parq_stats:
		for x in cell:
			x.border=thin_border

	range_stats_comp=act['A'+str(2*len(columns)+11) : 'E'+str(3*len(columns)+11)]
	for cell in range_stats_comp:
		for x in cell:
			x.border=thin_border

#Highlighting the Cell which is having mismatched Stats Comparison output "False"
	for i in range(1,500):
		for j in range(1,500):
			if (act.cell(row=i,column=j).value is False):
				act.cell(row=i,column=j).fill=PatterFill("solid",fgColor="fffff404")
			else:
				pass
else:
	pass

excel_handling.save('/tmp/data_validation.xlsx')


#-------------------------------------------------------------------STORING PROCESSED TABLE DETAILS FOR MAIL BODY-----------------------------------------------

processed_tble_detatils=[('','')]
processed_tble_detatils.append(tble_name,Target_Table)
schema=['Teradata Table Name', 'TargetTableName']

#Removing Emtpy Fields
tble_out=spark.createDataframe(processed_tble_detatils,schema=schema)
tble_out=tble_out.filter(tble_out.TargetTableName != '')

#Wrtie to CSV

tble_out.write.option("header","true")\
				.mode("append")\
				.csv('dbfs/teradata/tble.csv')